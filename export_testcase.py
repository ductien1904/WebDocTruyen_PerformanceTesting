"""
Script tự động xuất kết quả Locust ra file Excel test case.
Cách dùng:
  1. Chạy Locust với --csv: 
     python -m locust -f test_locust.py --host=http://localhost --csv=ket_qua --headless --users=10 --spawn-rate=2 --run-time=30s
  2. Sau đó chạy script này:
     python export_testcase.py
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import sys


# ──────────────────────────────────────────
# CẤU HÌNH
# ──────────────────────────────────────────
CSV_STATS   = "ket_qua_stats.csv"       # File CSV Locust tạo ra
CSV_HISTORY = "ket_qua_stats_history.csv"
OUTPUT_FILE = f"TestCase_Locust_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# Ngưỡng đánh giá PASS/FAIL
THRESHOLD_AVG_MS   = 2000   # Response time trung bình tối đa (ms)
THRESHOLD_P95_MS   = 5000   # Percentile 95 tối đa (ms)
THRESHOLD_FAIL_PCT = 1.0    # % lỗi tối đa cho phép


# ──────────────────────────────────────────
# MÀU SẮC
# ──────────────────────────────────────────
COLOR_HEADER     = "1F4E79"   # Xanh đậm
COLOR_PASS       = "C6EFCE"   # Xanh lá nhạt
COLOR_FAIL       = "FFC7CE"   # Đỏ nhạt
COLOR_WARN       = "FFEB9C"   # Vàng nhạt
COLOR_SUBHEADER  = "BDD7EE"   # Xanh nhạt
COLOR_WHITE      = "FFFFFF"


def border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def header_style(ws, cell_ref, value, bg=COLOR_HEADER, font_color="FFFFFF", bold=True, size=11):
    c = ws[cell_ref]
    c.value = value
    c.font = Font(name="Arial", bold=bold, color=font_color, size=size)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border()


def data_cell(ws, cell_ref, value, bg=COLOR_WHITE, bold=False, align="center", number_format=None):
    c = ws[cell_ref]
    c.value = value
    c.font = Font(name="Arial", bold=bold, size=10)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = border()
    if number_format:
        c.number_format = number_format


def load_csv():
    if not os.path.exists(CSV_STATS):
        print(f"❌ Không tìm thấy file '{CSV_STATS}'")
        print("   Hãy chạy Locust với flag --csv=ket_qua trước!")
        print("   Ví dụ:")
        print("   python -m locust -f test_locust.py --host=http://localhost --csv=ket_qua --headless --users=10 --spawn-rate=2 --run-time=30s")
        sys.exit(1)
    df = pd.read_csv(CSV_STATS)
    df = df[df["Name"] != "Aggregated"].copy()
    return df


def judge(row):
    reasons = []
    fail_pct = float(row.get("Failure Count", 0)) / max(float(row.get("Request Count", 1)), 1) * 100
    avg_ms   = float(row.get("Average Response Time", 0))
    p95_ms   = float(row.get("95%", 0))

    if fail_pct > THRESHOLD_FAIL_PCT:
        reasons.append(f"Lỗi {fail_pct:.1f}% > {THRESHOLD_FAIL_PCT}%")
    if avg_ms > THRESHOLD_AVG_MS:
        reasons.append(f"Avg {avg_ms:.0f}ms > {THRESHOLD_AVG_MS}ms")
    if p95_ms > THRESHOLD_P95_MS:
        reasons.append(f"P95 {p95_ms:.0f}ms > {THRESHOLD_P95_MS}ms")

    if reasons:
        return "FAIL", " | ".join(reasons)
    return "PASS", "Đạt tất cả ngưỡng"


# ──────────────────────────────────────────
# SHEET 1: TỔNG QUAN
# ──────────────────────────────────────────
def build_summary_sheet(wb, df):
    ws = wb.active
    ws.title = "Tổng Quan"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40

    ws.row_dimensions[1].height = 40
    header_style(ws, "A1", "BÁO CÁO TEST CASE - LOCUST", size=14)
    ws.merge_cells("A1:B1")

    info = [
        ("Thời gian xuất báo cáo", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("File CSV nguồn",          CSV_STATS),
        ("Tổng số endpoint test",   len(df)),
        ("Ngưỡng Avg Response",     f"≤ {THRESHOLD_AVG_MS} ms"),
        ("Ngưỡng P95 Response",     f"≤ {THRESHOLD_P95_MS} ms"),
        ("Ngưỡng % Lỗi",            f"≤ {THRESHOLD_FAIL_PCT}%"),
    ]

    for i, (k, v) in enumerate(info, start=3):
        ws.row_dimensions[i].height = 22
        data_cell(ws, f"A{i}", k, bg=COLOR_SUBHEADER, bold=True, align="left")
        data_cell(ws, f"B{i}", v, align="left")

    # Thống kê PASS/FAIL
    results = [judge(row) for _, row in df.iterrows()]
    n_pass = sum(1 for r, _ in results if r == "PASS")
    n_fail = len(results) - n_pass

    ws.row_dimensions[10].height = 22
    data_cell(ws, "A10", "Số test PASS", bg=COLOR_PASS, bold=True, align="left")
    data_cell(ws, "B10", n_pass, bg=COLOR_PASS)

    ws.row_dimensions[11].height = 22
    data_cell(ws, "A11", "Số test FAIL", bg=COLOR_FAIL, bold=True, align="left")
    data_cell(ws, "B11", n_fail, bg=COLOR_FAIL)

    ws.row_dimensions[12].height = 22
    overall_bg = COLOR_PASS if n_fail == 0 else COLOR_FAIL
    data_cell(ws, "A12", "Kết quả tổng thể", bg=overall_bg, bold=True, align="left")
    data_cell(ws, "B12", "✅ PASS" if n_fail == 0 else "❌ FAIL", bg=overall_bg, bold=True)


# ──────────────────────────────────────────
# SHEET 2: CHI TIẾT TEST CASE
# ──────────────────────────────────────────
def build_testcase_sheet(wb, df):
    ws = wb.create_sheet("Chi Tiết Test Case")

    cols = {
        "A": ("STT",              8),
        "B": ("Tên Endpoint",    35),
        "C": ("Phương thức",     12),
        "D": ("Số Request",      13),
        "E": ("Số Lỗi",          10),
        "F": ("% Lỗi",           10),
        "G": ("Avg (ms)",        12),
        "H": ("Min (ms)",        12),
        "I": ("Max (ms)",        12),
        "J": ("P50 (ms)",        12),
        "K": ("P90 (ms)",        12),
        "L": ("P95 (ms)",        12),
        "M": ("P99 (ms)",        12),
        "N": ("Req/s",           10),
        "O": ("Kết Quả",        12),
        "P": ("Ghi Chú",        40),
    }

    ws.row_dimensions[1].height = 35
    for col, (title, width) in cols.items():
        header_style(ws, f"{col}1", title)
        ws.column_dimensions[col].width = width

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        ws.row_dimensions[i].height = 22
        result, note = judge(row)
        bg = COLOR_PASS if result == "PASS" else COLOR_FAIL

        req_count  = int(row.get("Request Count", 0))
        fail_count = int(row.get("Failure Count", 0))
        fail_pct   = fail_count / max(req_count, 1) * 100

        values = [
            ("A", i - 1),
            ("B", row.get("Name", "")),
            ("C", row.get("Type", "GET")),
            ("D", req_count),
            ("E", fail_count),
            ("F", f"{fail_pct:.2f}%"),
            ("G", round(float(row.get("Average Response Time", 0)), 1)),
            ("H", round(float(row.get("Min Response Time", 0)), 1)),
            ("I", round(float(row.get("Max Response Time", 0)), 1)),
            ("J", round(float(row.get("50%", 0)), 1)),
            ("K", round(float(row.get("90%", 0)), 1)),
            ("L", round(float(row.get("95%", 0)), 1)),
            ("M", round(float(row.get("99%", 0)), 1)),
            ("N", round(float(row.get("Requests/s", 0)), 2)),
            ("O", f"✅ {result}" if result == "PASS" else f"❌ {result}"),
            ("P", note),
        ]

        for col, val in values:
            align = "left" if col in ("B", "P") else "center"
            cell_bg = bg if col in ("O", "P") else (COLOR_SUBHEADER if i % 2 == 0 else COLOR_WHITE)
            data_cell(ws, f"{col}{i}", val, bg=cell_bg, align=align)

    # Freeze header
    ws.freeze_panes = "A2"


# ──────────────────────────────────────────
# SHEET 3: BẢNG QUYẾT ĐỊNH
# ──────────────────────────────────────────
def build_decision_sheet(wb):
    ws = wb.create_sheet("Bảng Quyết Định")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 25

    ws.row_dimensions[1].height = 35
    ws.merge_cells("A1:E1")
    header_style(ws, "A1", "BẢNG QUYẾT ĐỊNH - ĐÁNH GIÁ KẾT QUẢ LOAD TEST", size=13)

    headers = ["Điều kiện", "% Lỗi", "Avg Response", "P95 Response", "Kết luận"]
    for i, h in enumerate(headers, start=1):
        col = get_column_letter(i)
        ws.row_dimensions[2].height = 28
        header_style(ws, f"{col}2", h, bg=COLOR_SUBHEADER, font_color="000000")

    rules = [
        ("Tất cả đều đạt ngưỡng",      f"≤ {THRESHOLD_FAIL_PCT}%",  f"≤ {THRESHOLD_AVG_MS}ms",  f"≤ {THRESHOLD_P95_MS}ms",  "✅ PASS"),
        ("Lỗi vượt ngưỡng",             f"> {THRESHOLD_FAIL_PCT}%",  "Bất kỳ",                   "Bất kỳ",                   "❌ FAIL"),
        ("Avg quá cao",                  "Đạt",                       f"> {THRESHOLD_AVG_MS}ms",  "Bất kỳ",                   "❌ FAIL"),
        ("P95 quá cao",                  "Đạt",                       "Đạt",                      f"> {THRESHOLD_P95_MS}ms",  "❌ FAIL"),
        ("Nhiều điều kiện vi phạm",      f"> {THRESHOLD_FAIL_PCT}%",  f"> {THRESHOLD_AVG_MS}ms",  f"> {THRESHOLD_P95_MS}ms",  "❌ FAIL (nghiêm trọng)"),
    ]

    for r, rule in enumerate(rules, start=3):
        ws.row_dimensions[r].height = 22
        bg = COLOR_PASS if "PASS" in rule[-1] else (COLOR_FAIL if "nghiêm trọng" in rule[-1] else COLOR_WARN)
        for c, val in enumerate(rule, start=1):
            col = get_column_letter(c)
            cell_bg = bg if c == 5 else (COLOR_WHITE if r % 2 == 0 else COLOR_SUBHEADER)
            data_cell(ws, f"{col}{r}", val, bg=cell_bg, align="center")


# ──────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────
def main():
    print("📊 Đang đọc kết quả Locust...")
    df = load_csv()
    print(f"   ✅ Tìm thấy {len(df)} endpoint")

    wb = Workbook()
    print("📝 Đang tạo file Excel...")
    build_summary_sheet(wb, df)
    build_testcase_sheet(wb, df)
    build_decision_sheet(wb)

    wb.save(OUTPUT_FILE)
    print(f"\n✅ Xuất thành công: {OUTPUT_FILE}")
    print("\n📌 Hướng dẫn sử dụng:")
    print("   Bước 1 - Chạy Locust với --csv:")
    print("   python -m locust -f test_locust.py --host=http://localhost --csv=ket_qua --headless --users=10 --spawn-rate=2 --run-time=30s")
    print("\n   Bước 2 - Xuất Excel:")
    print("   python export_testcase.py")


if __name__ == "__main__":
    main()
